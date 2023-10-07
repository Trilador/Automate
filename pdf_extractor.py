import PyPDF2
import contextlib
import os
import re
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import pickle
import hashlib

# ASCII Logo
ASCII_LOGO ="""


" ___  _          _ _ _            _     ___  _                _       
"|_ _|<_>._ _ _  | | | | ___  _ _ | |__ |  _>| |_  ___  ___  _| |_  ___
" | | | || ' ' | | | | |/ . \| '_>| / / | <__| . |/ ._><_> |  | |  <_-<
" |_| |_||_|_|_| |__/_/ \___/|_|  |_\_\ `___/|_|_|\___.<___|  |_|  /__/
"                                                                      

"""


# Constants for pickle paths
PICKLE_FOLDER_PATH = "pdf_folder_path.pkl"
PICKLE_OUTPUT_PATH = "DailyEmails_path.pkl"


def create_hash_from_row(row):
    """Create a unique hash for a row."""
    hash_object = hashlib.md5(str(row).encode())
    return hash_object.hexdigest()
    
def load_saved_paths():
    with contextlib.suppress(FileNotFoundError, EOFError, pickle.UnpicklingError):
        with open(PICKLE_FOLDER_PATH, 'rb') as f:
            folder_var.set(pickle.load(f))
    with contextlib.suppress(FileNotFoundError, EOFError, pickle.UnpicklingError):
        with open(PICKLE_OUTPUT_PATH, 'rb') as f:
            output_var.set(pickle.load(f))

def save_paths():
    with open(PICKLE_FOLDER_PATH, 'wb') as f:
        pickle.dump(folder_var.get(), f)
    with open(PICKLE_OUTPUT_PATH, 'wb') as f:
        pickle.dump(output_var.get(), f)

def extract_data():
    
    folder_path = folder_var.get()
    output_excel = output_var.get()
    if not folder_path or not output_excel:
        messagebox.showerror("Error", "Please select both the folder and the output Excel file.")
        return
    existing_hashes = set()
    if os.path.exists(output_excel):
        wb = load_workbook(output_excel)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.create_sheet("Data")
    else:
        wb = Workbook()
        ws = wb.create_sheet("Data")
    for file in os.listdir(folder_path):
        if file.endswith('.pdf'):
            with open(os.path.join(folder_path, file), 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                page = pdf_reader.pages[0]
                text = page.extract_text()
                pro_number = text[:7] if text[:7].isdigit() else "Not Found"
                fleet = text.split("NEW BERN - ")[1].split()[0]
                pickup_origin = next(
                    (line.split()[0] for line in text.split('\n') if ' OH ' in line),
                    "Not Found",
                )
                appt_date_time = "Not Found"
                if "Appointment" in text:
                    appt_dates = re.findall(r"Appointment (\d{2}/\d{2}/\d{2})", text)
                    appt_times = re.findall(r"@ (\d{2}:\d{2})", text)
                    if appt_dates and appt_times:
                        appt_date = appt_dates[-1]
                        appt_time = appt_times[-1]
                        appt_date_time = f"{appt_date} @ {appt_time}"
                reference_number = text.split("Ref # ")[1].split()[0]
                metadata = pdf_reader.metadata
                raw_date = metadata.get('/ModDate', 'Not Available')
                if parsed_date := re.search(r"D:(\d{4}\d{2}\d{2})", raw_date):
                    date_str = parsed_date[1]
                    date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                else:
                    date = 'Not Available'
                data_row = [date, fleet, pickup_origin, appt_date_time, reference_number, pro_number]
                data_hash = create_hash_from_row(data_row)
                if data_hash not in existing_hashes:
                    ws.append(data_row)
                    existing_hashes.add(data_hash)
                    print(f"Added data for {data_row}")
                else:
                    print(f"Data already exists for {data_row}")

    if output_excel.endswith('.xlsm'):
        wb.save(output_excel, keep_vba=True)
    else:
        wb.save(output_excel)
    messagebox.showinfo("Success", f"Data extracted and saved to {output_excel}")
    save_paths()
def select_output():
    output = filedialog.asksaveasfilename(defaultextension=".xlsx",
    filetypes=[("Excel files", "*.xlsx;*.xlsm"),
    ("Macro enabled Excel files", "*.xlsm"),
    ("All files", "*.*")])
    save_paths()


# Create the GUI window
root = tk.Tk()
root.title("Chep PDF extraction - Tim A")

# Display the ASCII Logo
logo_label = tk.Label(root, text=ASCII_LOGO, font=("Courier", 10))  # Use a monospace font for ASCII art
logo_label.pack(pady=20)

# Folder selection
folder_label = tk.Label(root, text="Select Folder:")
folder_label.pack(pady=10)

folder_var = tk.StringVar()
folder_entry = tk.Entry(root, textvariable=folder_var, width=70)  # Increased width
folder_entry.pack(pady=5)

def select_folder():
    folder = filedialog.askdirectory()
    folder_var.set(folder)

folder_button = tk.Button(root, text="Browse", command=select_folder)
folder_button.pack(pady=10)

# Output Excel file selection
output_label = tk.Label(root, text="Select Output Excel File:")
output_label.pack(pady=10)

output_var = tk.StringVar()
output_entry = tk.Entry(root, textvariable=output_var, width=70)  # Increased width
output_entry.pack(pady=5)

def select_output():
    output = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_var.set(output)

output_button = tk.Button(root, text="Browse", command=select_output)
output_button.pack(pady=10)

# Extract button
extract_button = tk.Button(root, text="Extract Data", command=extract_data)
extract_button.pack(pady=20)

# Load saved paths
load_saved_paths()

root.mainloop()